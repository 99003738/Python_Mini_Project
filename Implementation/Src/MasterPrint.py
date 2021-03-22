'''
importing the library file openpyxl as xl
'''
import openpyxl as xl
'''
importing UserInput.py file
'''
import UserInterface
'''
Importing the datafilter.py file
'''
import DataFilter

'''

Function name : writing_master()
Input : Global key search key
Output : after data found in respective sheet , copying the data to master sheet.

"""
'''
def writing_mastersheet():
    pathMasterWorkbook = UserInterface.outputPath[0]
    masterbook = xl.load_workbook(pathMasterWorkbook)
    lstsheet = masterbook.sheetnames
    length = len(lstsheet)
    currMassheet = masterbook.worksheets[0]
    print("\n")
    print("1. OverWrite ")
    print("2.New Sheet")
    choice = int(input("Enter Your choice : "))
    if choice == 1:
        currMassheet.delete_rows(2, currMassheet.max_row - 1)
        masterbook.save(UserInterface.outputPath[0])
    elif choice == 2:
        ws1 = masterbook.create_sheet("Mysheet")
        masterbook.save(UserInterface.outputPath[0])

'''
function name:: sending_data_master()
Input ::   pathvariable and searchitem
output :: if data matched then printing the data in master work book
'''


# single  path and global search item
def sending_data_master(pathvariable, searchitem = []):

    # creating workbook from where data to be searched
    workbook1 = xl.load_workbook(pathvariable)         # loading single path into the workbook1
    workbooksheet = workbook1.sheetnames               # counting the number of sheet present in the workbook and creating list of name
    lengthworkbook1 = len(workbooksheet)               # countin the sheets in the workbooksheet

    # taking path of master workbook to be printed .
    pathMasterWorkbook = UserInterface.outputPath[0]
    masterbook = xl.load_workbook(pathMasterWorkbook)     # taking the sheet of master workbook and intialising
    no_of_sheet = len(masterbook.sheetnames)
    # workSheetMaster = masterbook.active
    if no_of_sheet == 1:                                  # if there is only one sheet then taking only that one to print
        currMassheet = masterbook.worksheets[0]
    else:
        currMassheet = masterbook.worksheets[(no_of_sheet-1)]    # but if there are sheet present then add in the last location
    index = 0

    mastermaxrow = currMassheet.max_row
    mastermaxcol = currMassheet.max_column
    # currMassheet.cell(row= mastermaxrow+1, column=1).value = 'pathvariable'


    print(searchitem)
    for i in range(0, lengthworkbook1):                         # this loop is master loop running till length of the  workbook.
        sheets = workbook1.worksheets[i]                        # loading the first sheet in the sheets.
        wbMaxRow = sheets.max_row                               # finding the max row in the selected sheet
        wbMaxColumn = sheets.max_column                         # finding the length of the max column

                                                              # taking length of the searchitem
        lengthSearch = len(searchitem)
        colnum = 1

                                                          # finding the length of the master sheet
        mastermaxrow = currMassheet.max_row
        mastermaxcol = currMassheet.max_column

        rownum = mastermaxrow+2
        mastercol= 1
        # currMassheet.cell(row= rownum, column = mastercol).value = 'sheets'
        # rownum = rownum+1
        for colnum in range(4, wbMaxColumn + 1):             # this loop wil print the header of the selected sheet
            currMassheet.cell(row=rownum, column=mastercol).value = sheets.cell(row=1, column=colnum).value
            mastercol = mastercol+1
        mastercol= 1
        for i in range(2, wbMaxRow+1):                        # this loop srtart the  searching key search in the sheet
            data1 = searchitem[index]
            for j in range(1, 4):                              # this loop travel along the column to find the search key
                valuecheck = sheets.cell(row= i , column= j)
                value1 = str(valuecheck.value)
                # print(data1)
                # print(value1)
                if value1 == str(data1):
                    rownum = rownum + 1
                    mastercol =1
                    # print(data1)
                    # print(value1)
                    for k in range(4, wbMaxColumn+1):                # if the match key found then it will start printing into the sheet
                        cellprintpos = sheets.cell(row= i, column=k)
                        # print(cellprintpos)
                        cellprintval = cellprintpos.value
                        # print(cellprintval)
                        currMassheet.cell(row= rownum, column= mastercol).value =  sheets.cell(row=i, column= k).value
                        mastercol = mastercol+1
                    # index = index + 1
    masterbook.save(str(pathMasterWorkbook))                           # after doing the work saving the workbook

'''
this function is starting position of the program to run and calling the 
function as per finite state machine.
'''
obj1 = UserInterface.UserInterfaces()
obj2 = UserInterface.InputOutput()

obj1.user_selection()
obj2.choice_selection()
obj2.outputResult_path()

                                                    # asking for output result format
writing_mastersheet()                              # this will invoke the overwrite and new sheet creation section
length0fpathlist = len(UserInterface.pathList)         # counting the number of path entered by the user.
for d in range(0, length0fpathlist):               # running the loop till the whole path get executed
    path = UserInterface.pathList[d]                   # loading the first path in the path variable.
    check = DataFilter.validating_input(path)      # calling the data filter section for validating the data.
    if check == 1:                                      # if data validation is right then pass the path and input search key to the printing section of
        sending_data_master(path, UserInterface.inputList)
